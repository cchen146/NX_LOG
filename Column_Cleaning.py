from openpyxl.utils.datetime import from_excel
import datetime

error_dict = {}


class XlColumn:

    def __init__(self, worksheet, column, data_start_row=2, compulsory = False):
        self.worksheet = worksheet
        self.compulsory = compulsory
        self.column = column
        self.data_start_row = data_start_row

    def clean_data(self, row):
        return self.worksheet.cell(row=row, column=self.column).value

    def column_values(self):
        return [self.clean_data(row)
                for row in range(self.data_start_row, self.worksheet.max_row + 1)]


class DateColumn(XlColumn):

    def clean_data(self, row, *args, **kwargs):
        cell = self.worksheet.cell(row=row, column=self.column)
        if self.compulsory:
            if cell.internal_value == '':
                data = None
                try:
                    error_dict[row]
                except:
                    error_dict[row] = []
                error_dict[row].append(self.column)
                return data
            else:
                 try:
                     data = from_excel(cell.internal_value, cell.base_date)
                 except:
                     data = None
                     try:
                         error_dict[row]
                     except:
                         error_dict[row] = []
                     error_dict[row].append(self.column)
                 return data
        else:
            if str(cell.internal_value).strip() == '':
                data = None
                return data
            else:
                try:
                    if isinstance(cell.internal_value, str):
                        data = datetime.datetime.strptime(cell.internal_value, "%m/%d/%Y").date()
                    else:
                        data = from_excel(cell.internal_value, cell.base_date)
                except:
                    data = None
                    try:
                        error_dict[row]
                    except:
                        error_dict[row] = []
                    error_dict[row].append(self.column)
                return data


class TextColumn(XlColumn):

    def clean_data(self, row, *args, **kwargs):
        cell = self.worksheet.cell(row=row, column=self.column)
        if self.compulsory:
            if cell.internal_value == '':
                data = None
                try:
                    error_dict[row]
                except:
                    error_dict[row] = []
                error_dict[row].append(self.column)
                return data
            else:
                if cell.internal_value == None:
                    data = None
                    return data
                else:
                    data = str(cell.internal_value).replace('_x000D_','').strip()
                    return data
        else:
            if cell.internal_value == None:
                data = None
                return data
            else:
                data = str(cell.internal_value).replace('_x000D_', '').strip()
                return data


class IntColumn(XlColumn):

    def clean_data(self,row, *args, **kwargs):
        cell = self.worksheet.cell(row=row, column=self.column)
        if self.compulsory:
            if cell.internal_value == '':
                data = None
                try:
                    error_dict[row]
                except:
                    error_dict[row] = []
                error_dict[row].append(self.column)
                return data
            else:
                while True:
                    try:
                        data = int(str(cell.internal_value).replace('_x000D_', '').strip())
                    except:
                        data = None
                        try:
                            error_dict[row]
                        except:
                            error_dict[row] = []
                        error_dict[row].append(self.column)
                    return data
        else:
            if cell.internal_value == None:
                data = None
                return data
            else:
                while True:
                    try:
                        data = int(str(cell.internal_value).replace('_x000D_', '').strip())
                    except:
                        data = None
                        try:
                            error_dict[row]
                        except:
                            error_dict[row] = []
                        error_dict[row].append(self.column)
                    return data









