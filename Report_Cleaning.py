import ctypes  # An included library with Python install.
import csv
import datetime
import glob
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

import Column_Cleaning


def append_cleaned_columns(obj, column_data, worksheet, raw_data):
    for row in range(2, worksheet.max_row + 1):
        cell_value = obj.clean_data(row)
        column_data.append(cell_value)
    raw_data.append(column_data)

# construct a list of tuples; tuple contains rows of data in order of the columns in stg_table in database (stg_table_field)

def clean_data_by_column(worksheet, stg_table_field, input_column_position, int_column, date_column,
                         compulsory_field_name, key_fields):
    raw_data = []
    for field_name in stg_table_field:
        column_data = []
        if field_name in input_column_position.keys():
            col = input_column_position[field_name]
            if field_name in int_column:
                if field_name in key_fields:
                    compulsory = True
                else:
                    compulsory = False
                column_obj = Column_Cleaning.IntColumn(worksheet=worksheet, column=col, data_start_row=2,
                                                            compulsory=compulsory)
                append_cleaned_columns(obj=column_obj, column_data=column_data, worksheet=worksheet, raw_data=raw_data)
            elif field_name in date_column:
                if field_name in key_fields:
                    compulsory = True
                else:
                    compulsory = False
                column_obj = Column_Cleaning.DateColumn(worksheet=worksheet, column=col, data_start_row=2,
                                                             compulsory=compulsory)
                append_cleaned_columns(obj=column_obj, column_data=column_data, worksheet=worksheet, raw_data=raw_data)
            else:
                if field_name in key_fields:
                    compulsory = True
                else:
                    compulsory = False
                column_obj = Column_Cleaning.TextColumn(worksheet=worksheet, column=col, data_start_row=2,
                                                             compulsory=compulsory)
                append_cleaned_columns(obj=column_obj, column_data=column_data, worksheet=worksheet, raw_data=raw_data)
        elif field_name in compulsory_field_name:
            ctypes.windll.user32.MessageBoxW(0, "Missing [" + field_name + "] Column!", "ERROR", 1)
            exit()
        else:
            raw_data.append([])

    return raw_data

def convert_csv_to_xlsx(file):
    f = open(file, encoding='utf-8-sig')
    csv.register_dialect('colons', delimiter=',')
    reader = csv.reader(f, dialect='colons')
    wb = openpyxl.Workbook()
    today = datetime.date.today()
    dest_filename = r"C:\Users\ychen\Desktop\Testing\NX- FEDEX - Daily Report - {}.xlsx".format(today)
    ws = wb.worksheets[0]
    ws.title = "Forwarder_List"

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws.cell('%s%s' % (column_letter, (row_index + 1))).value = cell

    wb.save(filename=dest_filename)
    return dest_filename


def unify_input_file_ext(file):
    if file.lower().endswith('.csv'):
        file = convert_csv_to_xlsx(file)
        return file
    elif file.lower().endswith('.xlsx'):
        return file
    else:
        ctypes.windll.user32.MessageBoxW(0, 'file_extention is not acceptable', "Error", 1)


def log_exception(error_dict, input_column_position, fields_categories, fw, worksheet, today, error_log_file_path, sheet_name):
    # print out exception report
    if error_dict == {}:
        pass
    else:
        exception_log = openpyxl.Workbook()
        exception_ws = exception_log.create_sheet(sheet_name)
        del exception_log['Sheet']

        for col_name in input_column_position.keys():
            exception_ws.cell(row=1, column=input_column_position[col_name], value=col_name)

        exception_ws_row = 2

        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')

        for error_row in Column_Cleaning.error_dict.keys():
            error_row_list = [worksheet.cell(row=error_row, column=col).value for col in
                              range(1, worksheet.max_column + 1)]
            for exception_ws_column in range(1, worksheet.max_column + 1):
                cell = exception_ws.cell(row=exception_ws_row, column=exception_ws_column,
                                         value=error_row_list[exception_ws_column - 1])
                if exception_ws_column in Column_Cleaning.error_dict[error_row]:
                    cell.fill = redFill
            exception_ws_row += 1

            for date_column_name in fields_categories['date_column']:
                try:
                    date_column_position = input_column_position[date_column_name]
                    for row in range(1, exception_ws.max_row + 1):
                        exception_ws.cell(row=row, column=date_column_position).number_format = "MM/DD/YYYY"
                    exception_ws.column_dimensions[get_column_letter(date_column_position)].width = 13.75
                except:
                    pass

        exception_log.save(error_log_file_path + r'fw_daily_report_error_{}_{}.xlsx'.format(fw, today))
        ctypes.windll.user32.MessageBoxW(0, "{} daily report cleaned with errors".format(fw), "Error", 1)


def read_header_name_n_position(worksheet):
    input_column_position = {}  # Header
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        key = cell.value
        try:
            key = key.lower()
        except:
            print(key)
        value = col
        input_column_position[key] = value
    return input_column_position


def clean_report(file_name, fw, fields_categories, sheet_name):
    unify_input_file_ext(file_name)
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    worksheet = workbook[sheet_name]
    input_column_position = read_header_name_n_position(worksheet)
    today = datetime.date.today()
    Column_Cleaning.error_dict = {}
    raw_data = clean_data_by_column(worksheet=worksheet, input_column_position=input_column_position,
                                    **fields_categories)
    log_exception(error_dict = Column_Cleaning.error_dict, sheet_name = sheet_name, input_column_position = input_column_position, fields_categories = fields_categories, fw = fw, worksheet = worksheet, today = today, **kwargs)
    return raw_data


def get_latest_file(input_file, *args, **kwargs):
    input_file_name = max(glob.iglob(input_file), key=os.path.getctime)
    return input_file_name


def copy_file_without_last_line_csv(input_file, clean_file, *args, **kwargs):
    input_file_name = get_latest_file(input_file)
    f = open(input_file_name, encoding='utf-8')
    lines = f.readlines()
    lines = lines[:-1]
    f.close()
    with open(clean_file, 'w', encoding='utf-8') as nf:
        nf.writelines(lines)